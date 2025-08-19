import streamlit as st
import pandas as pd
import numpy as np
from scipy.interpolate import CubicSpline
import plotly.express as px
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, LineChart, Reference, Series

# --- Налаштування сторінки ---
st.set_page_config(page_title="Media Split Optimizer", layout="wide")
st.title("🎯 Media Split Optimizer — ТБ + Digital")

# --- Sidebar для параметрів ---
with st.sidebar:
    st.header("Параметри кампанії")
    budget = st.slider("Загальний бюджет (₴)", 100_000, 50_000_000, 5_000_000, step=100_000)
    flight_weeks = st.slider("Тривалість флайту (тижні)", 1, 30, 4)
    audience_size = st.number_input("Розмір аудиторії Digital (тис.)", min_value=1.0, value=1000.0)
    tv_cost_per_trp = st.number_input("Вартість 1 TRP ТБ (₴)", value=500.0)
    dig_cost_per_imp = st.number_input("Вартість 1 тис. імпресій Digital (₴)", value=5.0)
    tv_weekly_clutter = st.number_input("Конкурентний тиск ТБ (ТРП/тиждень)", value=150.0)
    dig_weekly_clutter = st.number_input("Конкурентний тиск Digital (TRP/тиждень)", value=300.0)
    split_step_percent = st.selectbox("Крок спліту (%)", [5, 10, 15, 20])
    n_options = st.slider("Кількість варіантів сплітів", 5, 15, 10)

# --- Введення TRP → Reach ---
with st.expander("TRP → Reach ТБ (5 точок)"):
    tv_trp_points, tv_reach_points = [], []
    for i in range(5):
        col1, col2 = st.columns(2)
        trp = col1.slider(f"TRP ТБ, точка {i+1}", 0.0, 500.0, float(i*50+50))
        reach = col2.slider(f"Reach_TV %, точка {i+1}", 0.0, 100.0, float(i*10+20))
        tv_trp_points.append(trp)
        tv_reach_points.append(reach/100)

with st.expander("TRP → Reach Digital (5 точок)"):
    dig_trp_points, dig_reach_points = [], []
    for i in range(5):
        col1, col2 = st.columns(2)
        trp = col1.slider(f"TRP Digital, точка {i+1}", 0.0, 500.0, float(i*10+10))
        reach = col2.slider(f"Reach_Digital %, точка {i+1}", 0.0, 100.0, float(i*5+10))
        dig_trp_points.append(trp)
        dig_reach_points.append(reach/100)

# --- Інтерполяція ---
tv_spline = CubicSpline(tv_trp_points, tv_reach_points)
dig_spline = CubicSpline(dig_trp_points, dig_reach_points)

# --- Генерація сплітів ---
split_step = split_step_percent / 100.0
split_values = np.arange(split_step, 1.0, split_step)
split_values = split_values[:n_options]

results = []
budget_warning = False
min_needed_budget = 0

for i, split in enumerate(split_values, start=1):
    tv_budget = budget * split
    dig_budget = budget * (1 - split)
    
    tv_trp = tv_budget / tv_cost_per_trp
    dig_imp = dig_budget / dig_cost_per_imp
    dig_trp = dig_imp / audience_size * 100
    
    tv_reach = float(np.clip(tv_spline(tv_trp), 0, 0.82))
    dig_reach = float(np.clip(dig_spline(dig_trp), 0, 0.99))
    cross_reach = tv_reach + dig_reach - tv_reach*dig_reach
    
    tv_weekly = tv_trp / flight_weeks
    dig_weekly = dig_trp / flight_weeks
    
    overall_ok = (tv_weekly >= tv_weekly_clutter) & (dig_weekly >= dig_weekly_clutter)
    if not overall_ok:
        budget_warning = True
        min_tv_budget = tv_weekly_clutter * flight_weeks * tv_cost_per_trp
        min_dig_budget = dig_weekly_clutter * flight_weeks * dig_cost_per_imp * audience_size / 100
        min_needed_budget = max(min_needed_budget, min_tv_budget + min_dig_budget)
    
    cpr = (tv_budget + dig_budget) / (cross_reach*100)
    cpt_dig = dig_budget / dig_trp
    
    results.append({
        "Опція": f"Опція {i}",
        "Спліт ТБ": f"{split*100:.0f}%",
        "Бюджет ТБ": int(tv_budget),
        "Бюджет Digital": int(dig_budget),
        "TRP_TV": round(tv_trp,1),
        "TRP_Digital": round(dig_trp,1),
        "Imp_Digital": round(dig_imp,1),
        "Reach_TV %": round(tv_reach*100,1),
        "Reach_Digital %": round(dig_reach*100,1),
        "Cross_Reach %": round(cross_reach*100,1),
        "Тиск ТБ/тижд": round(tv_weekly,1),
        "Тиск Digital/тижд": round(dig_weekly,1),
        "CPR": round(cpr,2),
        "CPT_Digital": round(cpt_dig,2),
        "Ефективний": overall_ok
    })

df = pd.DataFrame(results)
df['Доля ТБ %'] = df['Бюджет ТБ'] / (df['Бюджет ТБ'] + df['Бюджет Digital']) * 100
df['Доля Digital %'] = df['Бюджет Digital'] / (df['Бюджет ТБ'] + df['Бюджет Digital']) * 100

if budget_warning:
    st.warning(f"⚠️ Для досягнення конкурентного тиску потрібно щонайменше {int(min_needed_budget):,} ₴")

best_idx = df[df["Ефективний"]]["CPR"].idxmin() if df[df["Ефективний"]].shape[0]>0 else df["CPR"].idxmin()
best_option = df.loc[best_idx]

# --- KPI картки ---
st.subheader("🏆 Найкращий варіант")
col1, col2, col3 = st.columns(3)
col1.metric("Найнижчий CPR", f"{best_option['CPR']:.2f}")
col2.metric("Cross Reach %", f"{best_option['Cross_Reach %']:.1f}%")
col3.metric("TRP Digital", f"{best_option['TRP_Digital']:.1f}")

# --- Графіки ---
st.subheader("📊 Долі бюджету (stacked)")
df_budget_plot = df.melt(id_vars=["Опція"], value_vars=["Доля ТБ %", "Доля Digital %"],
                         var_name="Медіа", value_name="Доля %")
df_budget_plot["Медіа"] = df_budget_plot["Медіа"].replace({"Доля ТБ %": "ТБ", "Доля Digital %": "Digital"})
color_map = {"ТБ": "black", "Digital": "red"}

fig_budget = px.bar(df_budget_plot, x="Опція", y="Доля %", color="Медіа", text="Доля %",
                    title="Долі бюджету по медіа (stacked)", color_discrete_map=color_map)
fig_budget.update_yaxes(title_text="Доля бюджету (%)")
fig_budget.update_traces(texttemplate="%{text:.1f}%", textposition="inside")
st.plotly_chart(fig_budget, use_container_width=True)

st.subheader("📈 Охоплення по всіх опціях")
fig_reach = px.line(df, x="Опція", y=["Reach_TV %","Reach_Digital %","Cross_Reach %"],
                    markers=True, title="Reach TV / Digital / Cross")
st.plotly_chart(fig_reach, use_container_width=True)

# --- Таблиця ---
st.subheader("📋 Варіанти сплітів")
def highlight_rows(row):
    if row['Опція'] == best_option['Опція']:
        return ['background-color: lightblue']*len(row)
    elif row['Ефективний']:
        return ['background-color: lightgreen']*len(row)
    else:
        return ['background-color: salmon']*len(row)
st.dataframe(df.style.apply(highlight_rows, axis=1))

# --- Excel з графіками ---
st.subheader("⬇️ Завантаження Excel з графіками")
output = io.BytesIO()
wb = Workbook()
ws = wb.active
ws.title = "Splits"

for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Виділення найкращого CPR
cpr_col = None
for i, cell in enumerate(ws[1], start=1):
    if cell.value == "CPR":
        cpr_col = i
        break

if cpr_col:
    min_cpr = min(df["CPR"])
    fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
    for row in range(2, ws.max_row+1):
        if ws.cell(row=row, column=cpr_col).value == min_cpr:
            for col in range(1, ws.max_column+1):
                ws.cell(row=row, column=col).fill = fill

# --- Створення графіків у Excel ---
# 1. Stacked Bar Budget
bar_chart = BarChart()
bar_chart.title = "Долі бюджету по медіа"
bar_chart.type = "col"
bar_chart.style = 10
bar_chart.y_axis.title = "Доля %"
bar_chart.x_axis.title = "Опції"

data = Reference(ws, min_col=ws.max_column-1, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
ws.add_chart(bar_chart, "O2")

# 2. Line Chart Reach
line_chart = LineChart()
line_chart.title = "Reach TV / Digital / Cross"
line_chart.y_axis.title = "Reach %"
line_chart.x_axis.title = "Опції"
data2 = Reference(ws, min_col=8, min_row=1, max_col=10, max_row=ws.max_row)
line_chart.add_data(data2, titles_from_data=True)
line_chart.set_categories(cats)
ws.add_chart(line_chart, "O20")

# --- Скачування ---
wb.save(output)
output.seek(0)
st.download_button(
    label="Скачати Excel з графіками",
    data=output,
    file_name="media_split_with_charts.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


